import os
import json
import tempfile
from datetime import datetime
from typing import Optional
from flask import Flask, request, jsonify
from flask_cors import CORS
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

# ---------------------------
# Config & setup
# ---------------------------
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise RuntimeError("Missing OPENAI_API_KEY in environment")

client = OpenAI(api_key=OPENAI_API_KEY)

app = Flask(__name__)
CORS(app)

# ---------------------------
# Load grammar points
# ---------------------------
GRAMMAR_JSON_PATH = "grammar_points.json"

def load_grammar_points(path: str):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Could not find grammar file at: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

GRAMMAR_POINTS = load_grammar_points(GRAMMAR_JSON_PATH)

# ---------------------------
# Prompt builder
# ---------------------------
def grammar_points_block():
    return "\n".join([
        f"{p['id']}. {p['title']} - {p['rule']} Example: {p['example']}"
        for p in GRAMMAR_POINTS
    ])

def make_prompt(user_text: str, selected_id: Optional[int]) -> str:
    selected_block = ""
    if selected_id:
        selected_point = next((p for p in GRAMMAR_POINTS if p["id"] == int(selected_id)), None)
        if selected_point:
            selected_block = f"\nFocus on grammar point #{selected_point['id']}: {selected_point['title']}"

    return f"""
You are an ESL grammar coach. Analyze the learner’s sentence and reply ONLY in valid JSON with:
- corrected: string
- explanation: string (simple explanation for learner)
- grammar_ok: boolean
- score: integer 0-100
- matched_grammar_id: integer 1-70
- matched_grammar_label: string

Grammar Points:
{grammar_points_block()}
{selected_block}

Learner sentence:
\"\"\"{user_text}\"\"\"
"""

# ---------------------------
# Save learner logs (Excel)
# ---------------------------
LOG_FILE_XLSX = "learner_logs.xlsx"

def save_log(learner_id, data):
    row = [
        learner_id,
        datetime.now().isoformat(timespec="seconds"),
        data.get("transcript"),
        data.get("corrected"),
        data.get("explanation"),
        data.get("score"),
        data.get("matched_grammar_label"),
        data.get("selected_grammar_label"),
    ]

    if not os.path.exists(LOG_FILE_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "Logs"
        ws.append([
            "Learner ID",
            "Timestamp",
            "Transcript",
            "Correction",
            "Explanation",
            "Score",
            "Matched Grammar Point",
            "Selected Grammar Point",
        ])
        ws.append(row)
        wb.save(LOG_FILE_XLSX)
    else:
        wb = load_workbook(LOG_FILE_XLSX)
        ws = wb.active
        ws.append(row)
        wb.save(LOG_FILE_XLSX)

# ---------------------------
# Whisper transcription
# ---------------------------
def transcribe_audio_to_text(file_storage) -> str:
    if file_storage is None:
        return ""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".webm") as tmp:
            file_storage.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as audio_file:
            tr = client.audio.transcriptions.create(
                model="whisper-1",
                file=audio_file
            )
        return tr.text.strip() if hasattr(tr, "text") else ""
    except Exception as e:
        print("Transcription error:", e)
        return ""

# ---------------------------
# GPT grammar check
# ---------------------------
def run_grammar_llm(user_text: str, grammar_id: Optional[int]) -> dict:
    prompt = make_prompt(user_text, grammar_id)
    completion = client.chat.completions.create(
        model="gpt-4o-mini",
        temperature=0,
        messages=[
            {"role": "system", "content": "You are a JSON-only ESL grammar evaluator. Always output ONLY valid JSON."},
            {"role": "user", "content": prompt}
        ]
    )
    raw = completion.choices[0].message.content.strip()
    try:
        result = json.loads(raw)
    except Exception:
        start = raw.find("{")
        end = raw.rfind("}")
        if start != -1 and end != -1:
            try:
                result = json.loads(raw[start:end+1])
            except Exception:
                result = {"error": "Could not parse GPT response", "raw": raw}
        else:
            result = {"error": "Invalid response", "raw": raw}
    return result

# ---------------------------
# Routes
# ---------------------------
@app.route("/")
def home():
    return "Grammar checker backend running"

@app.route("/api/grammar-points", methods=["GET"])
def api_grammar_points():
    return jsonify(GRAMMAR_POINTS)

@app.route("/api/text", methods=["POST"])
def api_text():
    typed = request.form.get("typed", "").strip()
    learner_id = request.form.get("learner_id", "anonymous")
    grammar_id = request.form.get("grammar_id", "").strip()

    if not typed:
        return jsonify({"error": "No text provided"}), 400

    sel_id = int(grammar_id) if grammar_id.isdigit() else None
    out = run_grammar_llm(typed, sel_id)
    out["transcript"] = typed
    out["selected_grammar_label"] = next(
        (p["title"] for p in GRAMMAR_POINTS if str(p["id"]) == grammar_id),
        None
    )
    save_log(learner_id, out)
    return jsonify(out)

@app.route("/api/grammar", methods=["POST"])
def api_grammar():
    learner_id = request.form.get("learner_id", "anonymous")
    grammar_id = request.form.get("grammar_id", "").strip()
    typed = request.form.get("typed", "").strip()
    transcript = typed

    if not typed:
        file = request.files.get("audio")
        transcript = transcribe_audio_to_text(file)

    if not transcript:
        return jsonify({"error": "No speech or text found."}), 400

    sel_id = int(grammar_id) if grammar_id.isdigit() else None
    out = run_grammar_llm(transcript, sel_id)
    out["transcript"] = transcript
    out["selected_grammar_label"] = next(
        (p["title"] for p in GRAMMAR_POINTS if str(p["id"]) == grammar_id),
        None
    )
    save_log(learner_id, out)
    return jsonify(out)

# ---------------------------
# Render-compatible start
# ---------------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
